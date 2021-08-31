'''
Author: NEFU AB_IN
Date: 2021-08-31 15:31:15
FilePath: \Vscode\Excel&python\example.py
LastEditTime: 2021-08-31 15:35:34
'''
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
# https://openpyxl.readthedocs.io/en/stable/styles.html

data = [
    {
        'name': 'A',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': 'B',
        'tall': 177,
        'age': 28,
        'weight': 90
    },
    {
        'name': 'C',
        'tall': 160,
        'age': 30,
        'weight': 60
    },
    {
        'name': 'D',
        'tall': 155,
        'age': 50,
        'weight': 50
    },
    {
        'name': 'E',
        'tall': 170,
        'age': 46,
        'weight': 99
    }
]

wb = Workbook()
ws = wb.active

title = ['name', 'height', 'old', 'weight']
ws.append(title)

for person in data:
    ws.append(list(person.values()))

for col in range(2,5):
    char = get_column_letter(col)
    ws[char + '7'] = f'=AVERAGE({char}2:{char}6)'# 直接输入公式

for col in range(1,5):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color="000000FF") #粗体

wb.save('data.xlsx')

