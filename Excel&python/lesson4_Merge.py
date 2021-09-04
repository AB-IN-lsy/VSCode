'''
Author: NEFU AB_IN
Date: 2021-08-31 13:14:34
FilePath: \Vscode\Excel&python\lesson4_Merge.py
LastEditTime: 2021-09-02 19:37:51
'''
from openpyxl import Workbook, load_workbook

wb = load_workbook('new_excel.xlsx')
ws = wb.active
ws.merge_cells('A1:E1') #合并单元格
ws.unmerge_cells('A1:E1') #还原，但原本的资料不会被恢复

wb.save('new_excel.xlsx')

