'''
Author: NEFU AB_IN
Date: 2021-08-31 12:20:51
FilePath: \Vscode\Excel&python\lesson1_Load.py
LastEditTime: 2021-08-31 12:47:47
'''
from openpyxl import workbook, load_workbook
#适用于2010以上excel

wb = load_workbook('test.xlsx') # 读取excel档案
ws = wb.active # 选取工作表(这里选取的是默认的)
print(ws)
#<Worksheet "Sheet1">
print(ws['A5'])
#<Cell 'Sheet1'.A5>
print(ws['A5'].value) #取得表格内容Cell
#NEFU
ws['C12'].value = '2019级' #修改
#-------------------------------------------------
print(wb.sheetnames)
#['Sheet1', 'Sheet2', 'Sheet3']
#-------------------------------------------------
ws = wb['Sheet2'] #自定义选择工作表
print(ws)
#<Worksheet "Sheet2">
#-------------------------------------------------
wb.create_sheet('lsy') #创建工作表
print(wb.sheetnames)
#['Sheet1', 'Sheet2', 'Sheet3', 'lsy']
#-------------------------------------------------


wb.save('test.xlsx') #保存修改才会生效
