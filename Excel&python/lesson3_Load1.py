'''
Author: NEFU AB_IN
Date: 2021-08-31 12:54:20
FilePath: \Vscode\Excel&python\lesson3_Load1.py
LastEditTime: 2021-08-31 13:09:37
'''
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('new_excel.xlsx')
ws = wb.active

for row in range(1, 5): #必须以1作为起始
    for col in range(1, 5):
        col_char = get_column_letter(col) # 比如 1->A
        print(ws[col_char + str(row)].value, end = " ")
    print()
wb.save('new_excel.xlsx')