'''
Author: NEFU AB_IN
Date: 2021-08-31 12:46:39
FilePath: \Vscode\Excel&python\lesson2_Create.py
LastEditTime: 2021-08-31 12:53:05
'''
from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = 'lsy' #命名工作表
ws['A1'].value = 1

ws.append([123, 456, 789, 0]) #新增一横排的列表，不会覆盖
ws.append([123, 456, 789, 0]) 
ws.append([123, 456, 789, 0]) 

wb.save('new_excel.xlsx') #如果存在就会覆盖